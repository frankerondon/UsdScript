"""
Functional script version with XLSX Files
"""


import os 
#import sys
import requests
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

#Ejecucion 10:00 AM && 02:00 PM

# Define the API endpoint
BASE_URL = "https://pydolarve.org/"
ENDPOINT = "api/v2/dollar"
API_URL = BASE_URL + ENDPOINT

#Define CSV file 
XLSX_FILENAME = "Precio Dolar.xlsx"
XLSX_HEADERS = ['FECHA', 'HORA', 'BCV', 'PARALELO', 'PROMEDIO']


def fetch_dollar_data():
    """
    Fetches dollar exchange rate data from the pydolarve API
    and stores it in a variable.
    """
    try:
        # Make the GET request to the API
        response = requests.get(API_URL)

        # Raise an exception for bad status codes (4xx or 5xx)
        response.raise_for_status()

        # Parse the JSON response directly using response.json()
        # This will convert the JSON string into a Python dictionary or list
        dollar_data = response.json()
        """
        # Now, dollar_data holds the JSON content as a Python object
        #print("Successfully fetched and parsed data:")
        #print(dollar_data) # You can inspect the data here
        """
        return dollar_data

    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
    except requests.exceptions.ConnectionError as conn_err:
        print(f"Error Connecting: {conn_err}")
    except requests.exceptions.Timeout as timeout_err:
        print(f"Timeout Error: {timeout_err}")
    except requests.exceptions.RequestException as req_err:
        print(f"OOps: Something Else went wrong with the request: {req_err}")
    except json.JSONDecodeError:
        print("Error: Failed to decode JSON from the response.")
    return None


def format_date_spanish(date_parts):
    """
    Converts a list of date parts [day, month_name, year]
    to "DD/MM/YYYY" string format.
    Example: ['29', 'mayo', '2025'] -> "29/05/2025"
    """
    if not date_parts or len(date_parts) != 3:
        return "Invalid date parts"

    day, month_name, year = date_parts

    month_map_spanish = {
        "enero": "01", 
        "febrero": "02", 
        "marzo": "03", 
        "abril": "04",
        "mayo": "05", 
        "junio": "06", 
        "julio": "07", 
        "agosto": "08",
        "septiembre": "09", 
        "octubre": "10", 
        "noviembre": "11", 
        "diciembre": "12"
    }

    month_number = month_map_spanish.get(month_name.lower()) # Use lower() for case-insensitivity

    if not month_number:
        return f"Unknown month name: {month_name}"

    # Ensure day is two digits (though it usually will be from your split)
    day_formatted = day.zfill(2)

    return f"{day_formatted}/{month_number}/{year}"



def get_formatted_date_from_api(api_data):
    """
    Extracts the date from the API response, processes it,
    and returns it in DD/MM/YYYY format.
    Returns a formatted date string or a default/error message string if issues occur.
    """
    if not api_data:
        print("Error: No API data provided to get_formatted_date_from_api.")
        return "Date unavailable (no API data)"

    date_data = api_data.get("datetime")
    if not date_data:
        print("Error: 'datetime' key not found in API data.")
        return "Date unavailable (datetime missing)"

    raw_date_api = date_data.get("date") # e.g., "jueves, 29 de mayo de 2025"
    if not raw_date_api:
        print("Error: 'date' key not found in datetime data.")
        return "Date unavailable (date string missing)"

    try:
        # Get " 29 de mayo de 2025", then split by "de"
        unstripped_date_parts = raw_date_api.split(",")[1].split("de")
        # Remove leading/trailing whitespace from each part
        # Result: ['29', 'mayo', '2025']
        cleaned_date_parts = [part.strip() for part in unstripped_date_parts]

        # Convert to DD/MM/YYYY format
        formatted_date = format_date_spanish(cleaned_date_parts)
        return formatted_date
        
    except IndexError:
        print(f"Error: Date string '{raw_date_api}' format was not as expected. Could not split correctly.")
        return "Date unavailable (format error)"
    
    except Exception as e:
        print(f"An unexpected error occurred during date processing: {e}")
        return "Date unavailable (processing error)"


def getHourAlCambio(api_data):
    strHour = api_data.get("last_update").split(",")[1].strip()
    #print(f"{strHour}")
    return strHour
    

def exchangesPrice(api_data):
    usdCurrPrice = api_data.get("price")
    return usdCurrPrice
    #print(f"{usdCurrPrice}")


def getAveragePrice(bcv, par):
    result = round((bcv + par) / 2, 2)
    #print(f"Average Price: {result}")
    return result


def apply_excel_styling(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Define el ancho de las columnas en unidades de caracteres de Excel.
    # La conversión de píxeles a unidades de ancho de caracteres es aproximada y
    # depende de la fuente y el tamaño predeterminados. Un factor común es ~7.5-8 píxeles por unidad.
    pixel_to_char_width_factor = 7.5 
    
    column_settings = {
        'A': {'width_pixels': 99},  # FECHA
        'B': {'width_pixels': 94},  # HORA
        'C': {'width_pixels': 78},  # BCV
        'D': {'width_pixels': 88},  # PARALELO
        'E': {'width_pixels': 84}   # PROMEDIO
    }

    for col_letter, settings in column_settings.items():
        sheet.column_dimensions[col_letter].width = settings['width_pixels'] / pixel_to_char_width_factor

    # Define el estilo de borde
    thin_border_side = Side(border_style="thin", color="000000")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Aplica bordes a todas las filas de datos (el encabezado es la fila 1, los datos comienzan en la fila 2)
    if sheet.max_row > 1: # Solo aplicar si hay filas de datos
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(XLSX_HEADERS)):
            for cell in row:
                cell.border = cell_border
    
    workbook.save(filename)




if __name__ == "__main__":

    # Execute the function and store the result
    # Get Whole data from API 
    data_from_api = fetch_dollar_data()

    sources_data = data_from_api.get("monitors") 
    alcambio_data = sources_data.get("alcambio")
    bcv_data = sources_data.get("bcv")
    enparalelo_data = sources_data.get("enparalelovzla")
    
    bcvPrice = exchangesPrice(bcv_data)
    parPrice = exchangesPrice(alcambio_data)

    toExcelFile = [get_formatted_date_from_api(data_from_api), getHourAlCambio(alcambio_data), bcvPrice, parPrice, getAveragePrice(bcvPrice, parPrice)]

    print(f"Data prepared for Excel: {toExcelFile}")
    
    # --- Writing to XLSX using pandas ---
    # If the file Precio Dolar.xlsx is opened in Excel, it will raise an error when trying to write to it.

    try:
        # Create a DataFrame from the toExcelFile list
        # Since toCSVFile is a single row, it needs to be a list of lists for the DataFrame constructor
        df_new_row = pd.DataFrame([toExcelFile], columns=XLSX_HEADERS)

        file_exists = os.path.isfile(XLSX_FILENAME)

        if not file_exists:
            # File does not exist, write with header
            df_new_row.to_excel(XLSX_FILENAME, index=False, header=True, engine='openpyxl')
            print(f"Excel file '{XLSX_FILENAME}' created with headers and data.")
        
        else:
        
            # File exists, read it, append new data, and write back
            try:
                df_existing = pd.read_excel(XLSX_FILENAME, engine='openpyxl')
                df_combined = pd.concat([df_existing, df_new_row], ignore_index=True)
                df_combined.to_excel(XLSX_FILENAME, index=False, header=True, engine='openpyxl')

            except FileNotFoundError: 
                # Should not happen if file_exists is true, but as a safeguard
                df_new_row.to_excel(XLSX_FILENAME, index=False, header=True, engine='openpyxl')
                print(f"Excel file '{XLSX_FILENAME}' was not found despite check, created new.")            
            
            print(f"Data appended to existing Excel file '{XLSX_FILENAME}'.")

        # Apply styling after data is written
        apply_excel_styling(XLSX_FILENAME)

        except Exception as e:
            print(f"An error occurred during Excel writing or styling with pandas: {e}")
